import streamlit as st
import pandas as pd
import os
from io import BytesIO
from openpyxl import Workbook
from pdf2image import convert_from_bytes
from processor import process_pdf, KIF_HEADERS

def get_app_password():
    try:
        pw = st.secrets.get("APP_PASSWORD", "")
        if pw:
            return pw
    except Exception:
        pass
    return os.environ.get("APP_PASSWORD", "")

APP_PASSWORD = get_app_password()

# ── Session state init ──
_params = st.query_params
if "authenticated" not in st.session_state:
    st.session_state.authenticated = _params.get("auth") == "1"
if "page" not in st.session_state:
    st.session_state.page = _params.get("page", "home")

# ── Page config (must be first Streamlit command) ──
_layout = "wide" if st.session_state.page == "kif" else "centered"
st.set_page_config(page_title="BS BIRO", page_icon="📄", layout=_layout)

# ── Helpers ──
def get_logo_b64():
    logo_path = os.path.join(os.path.dirname(__file__), "images", "logo.png")
    if os.path.exists(logo_path):
        import base64
        return base64.b64encode(open(logo_path, "rb").read()).decode()
    return None

def get_api_key():
    try:
        key = st.secrets.get("OPENAI_API_KEY", "")
        if key:
            return key
    except Exception:
        pass
    return os.environ.get("OPENAI_API_KEY", "")

logo_b64 = get_logo_b64()

# ═══════════════════════════════════════════
# HOME PAGE
# ═══════════════════════════════════════════
if st.session_state.page == "home":

    st.markdown("""
    <style>
    header {visibility:hidden;}
    #MainMenu {visibility:hidden;}
    footer {visibility:hidden;}
    .stApp {background-color:#f6d9c0;}
    .block-container {max-width:600px; padding-top:2rem;}
    .logo-row {display:flex; align-items:center; gap:14px; margin-bottom:4px; justify-content:center;}
    .logo-row img {height:64px; width:auto;}
    .logo-row .app-title {font-size:2.4rem; font-weight:700; margin:0;}
    button[kind="primary"] {
        background:#0e8a3e !important; color:white !important; border:none !important;
    }
    button[kind="primary"]:hover {
        background:#0b6e31 !important; color:white !important;
    }
    .copyright {text-align:center; font-size:11px; color:#94a3b8; margin-top:30px;}
    </style>
    """, unsafe_allow_html=True)

    if logo_b64:
        st.markdown(f'<div class="logo-row"><img src="data:image/png;base64,{logo_b64}" /><div class="app-title">BS BIRO</div></div>', unsafe_allow_html=True)
    else:
        st.markdown('<div class="logo-row"><div class="app-title">BS BIRO</div></div>', unsafe_allow_html=True)

    st.markdown('<p style="text-align:center; color:#64748b; margin-top:4px;">Automatska obrada PDF računa</p>', unsafe_allow_html=True)

    if not st.session_state.authenticated:
        st.markdown("---")
        st.subheader("Prijava")
        password = st.text_input("Unesi šifru", type="password", placeholder="Šifra...")
        if st.button("Prijavi se", type="primary", use_container_width=True):
            if password == APP_PASSWORD:
                st.session_state.authenticated = True
                st.query_params["auth"] = "1"
                st.rerun()
            else:
                st.error("Pogrešna šifra. Pokušaj ponovo.")
        st.markdown('<div class="copyright">Sva prava zadržana, Amir Basic( basic.amir99@gmail.com )</div>', unsafe_allow_html=True)
        st.stop()

    # Authenticated - show navigation
    st.markdown("---")
    st.subheader("Odaberi modul")

    col1, col2 = st.columns(2)
    with col1:
        if st.button("📤 KIF — Knjiga Izlaznih Faktura", use_container_width=True):
            st.session_state.page = "kif"
            st.query_params["page"] = "kif"
            st.rerun()
        st.caption("Obrada izlaznih računa koje tvoja firma izdaje kupcima.")
    with col2:
        if st.button("📥 KUF — Knjiga Ulaznih Faktura", use_container_width=True):
            st.session_state.page = "kuf"
            st.query_params["page"] = "kuf"
            st.rerun()
        st.caption("Obrada ulaznih računa koje tvoja firma prima od dobavljača.")

    st.markdown('<div class="copyright">Sva prava zadržana, Amir Basic( basic.amir99@gmail.com )</div>', unsafe_allow_html=True)

# ═══════════════════════════════════════════
# KIF PAGE
# ═══════════════════════════════════════════
elif st.session_state.page == "kif":

    st.markdown("""
    <style>
    header {visibility:hidden;}
    #MainMenu {visibility:hidden;}
    footer {visibility:hidden;}
    .block-container {max-width:100%; padding-top:1rem; padding-bottom:2rem; padding-left:2rem; padding-right:2rem;}
    .stApp {background-color:#f6d9c0;}
    [data-testid="stDataEditor"] {min-height:600px;}
    .logo-row {display:flex; align-items:center; gap:14px; margin-bottom:4px;}
    .logo-row img {height:52px; width:auto;}
    .logo-row .app-title {font-size:2rem; font-weight:700; margin:0;}
    .steps {background:#f8fafc; border:1px solid #e2e8f0; border-radius:10px; padding:14px 18px; margin:10px 0 16px 0; font-size:13.5px; line-height:1.7; color:#334155;}
    .steps .copyright {margin-top:12px; padding-top:10px; border-top:1px solid #e2e8f0; font-size:11px; color:#94a3b8;}
    div.stDownloadButton > button {background:#0e8a3e; color:white; border:none; border-radius:8px; font-weight:500;}
    div.stDownloadButton > button:hover {background:#0b6e31; color:white;}
    #pdf_download > button {background:#18181b !important; color:white !important;}
    #pdf_download > button:hover {background:#3f3f46 !important; color:white !important;}
    button[kind="primary"] {background:#0e8a3e !important; color:white !important; border:none !important;}
    button[kind="primary"]:hover {background:#0b6e31 !important; color:white !important;}
    </style>
    """, unsafe_allow_html=True)

    # Header
    top_left, top_right = st.columns([3, 2])

    with top_left:
        if st.button("← Nazad", key="back_home"):
            st.session_state.page = "home"
            st.query_params["page"] = "home"
            st.rerun()

        if logo_b64:
            st.markdown(f'<div class="logo-row"><img src="data:image/png;base64,{logo_b64}" /><div class="app-title">KIF — BS BIRO</div></div>', unsafe_allow_html=True)
        else:
            st.markdown('<div class="logo-row"><div class="app-title">KIF — BS BIRO</div></div>', unsafe_allow_html=True)

        st.caption("Knjiga Izlaznih Faktura")
        st.markdown("""
        <div class="steps">
            <strong>Kako koristiti:</strong><br>
            1. Upload-uj jedan ili više PDF računa (drag & drop ili klikni Browse)<br>
            2. Klikni <b>Obradi račune</b> — AI automatski izvlači podatke iz svakog PDF-a<br>
            3. Pregledaj i edituj podatke u tabeli ako treba<br>
            4. Preuzmi gotov Excel ili CSV fajl
            <div class="copyright">Sva prava zadržana, Amir Basic( basic.amir99@gmail.com )</div>
        </div>
        """, unsafe_allow_html=True)

        uploaded_files = st.file_uploader("Prevuci ili odaberi PDF račune", type=["pdf"], accept_multiple_files=True)

    with top_left:
        if not uploaded_files:
            st.info("Dodaj račune za početak obrade.")
            st.stop()

        st.write(f"**{len(uploaded_files)}** račun(a) odabrano")

    # Session state
    if "results" not in st.session_state:
        st.session_state.results = []
    if "logs" not in st.session_state:
        st.session_state.logs = []
    if "pdf_map" not in st.session_state:
        st.session_state.pdf_map = {}

    with top_left:
        process_clicked = st.button("Obradi račune", type="primary", use_container_width=True)

    if process_clicked:
        api_key = get_api_key()
        with top_left:
            if not api_key:
                st.error("OpenAI API ključ nije pronađen. Dodaj ga u .streamlit/secrets.toml ili .env")
                st.stop()

        st.session_state.results = []
        st.session_state.logs = []
        st.session_state.pdf_map = {}
        seen = set()

        with top_left:
            with st.spinner("AI obrađuje račune, molimo sačekajte..."):
                progress = st.progress(0, text="Pokrećem obradu...")
                for i, file in enumerate(uploaded_files):
                    progress.progress(i / len(uploaded_files), text=f"Obrađujem {i+1}/{len(uploaded_files)}: {file.name}")
                    try:
                        pdf_bytes = file.read()
                        data = process_pdf(pdf_bytes, filename=file.name, api_key=api_key)
                        broj = data.get("BRDOKFAKT", "")
                        if broj and broj in seen:
                            st.session_state.logs.append(("warn", f"{file.name} — duplikat računa {broj}"))
                            continue
                        seen.add(broj)
                        idx = len(st.session_state.results)
                        st.session_state.results.append(data)
                        st.session_state.pdf_map[idx] = pdf_bytes
                        st.session_state.logs.append(("ok", f"{file.name} — {data.get('NAZIVPP','?')} — {data.get('IZNAKFT','?')} KM"))
                    except Exception as e:
                        st.session_state.logs.append(("err", f"{file.name} — {str(e)}"))
                    progress.progress((i + 1) / len(uploaded_files))
                progress.progress(1.0, text=f"Gotovo! Obrađeno {len(st.session_state.results)} račun(a)")

    # Results
    if st.session_state.results:
        with top_left:
            st.divider()
            for t, msg in st.session_state.logs:
                if t == "ok":
                    st.success(msg, icon="✅")
                elif t == "err":
                    st.error(msg, icon="❌")
                else:
                    st.warning(msg, icon="⚠️")

            st.divider()
            st.subheader("Podaci")
            st.caption("Klikni na polje u tabeli da edituješ prije downloada")

            df = pd.DataFrame(st.session_state.results, columns=KIF_HEADERS)
            for col in KIF_HEADERS:
                if col not in df.columns:
                    df[col] = ""

            edited_df = st.data_editor(df[KIF_HEADERS], use_container_width=True, hide_index=True, num_rows="dynamic", key="data_editor")

            def create_excel(dataframe):
                output = BytesIO()
                wb = Workbook()
                ws = wb.active
                ws.title = "Racuni"
                for c, h in enumerate(KIF_HEADERS, 1):
                    ws.cell(row=1, column=c, value=h)
                for r, row in dataframe.iterrows():
                    for c, h in enumerate(KIF_HEADERS, 1):
                        ws.cell(row=r + 2, column=c, value=row.get(h, ""))
                wb.save(output)
                return output.getvalue()

            st.divider()
            e1, e2 = st.columns(2)
            with e1:
                st.download_button("Preuzmi Excel", create_excel(edited_df), "racuni.xlsx", type="primary", use_container_width=True)
            with e2:
                st.download_button("Preuzmi CSV", edited_df.to_csv(index=False, sep=";", encoding="utf-8-sig"), "racuni.csv", use_container_width=True)

        with top_right:
            st.subheader("PDF pregled")
            selected = st.selectbox(
                "Odaberi račun za pregled",
                options=range(len(st.session_state.results)),
                format_func=lambda i: f"{st.session_state.results[i].get('NAZIVPP','?')} — {st.session_state.results[i].get('BRDOKFAKT','')}",
            )
            pdf_bytes = st.session_state.pdf_map.get(selected)
            if pdf_bytes:
                st.download_button("Preuzmi ovaj PDF", pdf_bytes, "racun.pdf", use_container_width=True, key="pdf_download")
                pages = convert_from_bytes(pdf_bytes, dpi=150)
                for page in pages:
                    st.image(page, use_container_width=True)

# ═══════════════════════════════════════════
# KUF PAGE
# ═══════════════════════════════════════════
elif st.session_state.page == "kuf":

    st.markdown("""
    <style>
    header {visibility:hidden;}
    #MainMenu {visibility:hidden;}
    footer {visibility:hidden;}
    .stApp {background-color:#f6d9c0;}
    .block-container {max-width:600px; padding-top:2rem;}
    .logo-row {display:flex; align-items:center; gap:14px; margin-bottom:4px; justify-content:center;}
    .logo-row img {height:52px; width:auto;}
    .logo-row .app-title {font-size:2rem; font-weight:700; margin:0;}
    .copyright {text-align:center; font-size:11px; color:#94a3b8; margin-top:30px;}
    </style>
    """, unsafe_allow_html=True)

    if st.button("← Nazad"):
        st.session_state.page = "home"
        st.query_params["page"] = "home"
        st.rerun()

    if logo_b64:
        st.markdown(f'<div class="logo-row"><img src="data:image/png;base64,{logo_b64}" /><div class="app-title">KUF — BS BIRO</div></div>', unsafe_allow_html=True)
    else:
        st.markdown('<div class="logo-row"><div class="app-title">KUF — BS BIRO</div></div>', unsafe_allow_html=True)

    st.caption("Knjiga Ulaznih Faktura")
    st.markdown("---")
    st.info("KUF modul je trenutno u izradi. Ova funkcionalnost će biti dostupna uskoro.")
    st.markdown('<div class="copyright">Sva prava zadržana, Amir Basic( basic.amir99@gmail.com )</div>', unsafe_allow_html=True)

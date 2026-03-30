import openai
import base64
import json
import random
import re
import fitz
from openpyxl import load_workbook
from pdf2image import convert_from_path
from io import BytesIO
import tempfile
import os
import time
from concurrent.futures import ThreadPoolExecutor, as_completed

MIN_TEXT_LENGTH = 100

KIF_HEADERS = [
    "REDBR", "TIPDOK", "BRDOKFAKT", "DATUMF",
    "NAZIVPP", "SJEDISTEPP", "IDDVPP", "JIBPUPP",
    "IZNAKFT", "IZNOSNOV", "IZNPDV", "REF", "OSL", "KONTO",
]

POZNATI_PARTNERI = [
    #todo
]


_CP1250_FIX = {
    '\u00C6': 'Ć', '\u00E6': 'ć',
    '\u00C8': 'Č', '\u00E8': 'č',
    '\u00D0': 'Đ', '\u00F0': 'đ',
}


def _fix_cp1250(s):
    for bad, good in _CP1250_FIX.items():
        s = s.replace(bad, good)
    return s


def load_kupci_names(xlsx_path="kupci.xlsx"):
    """Učitava listu pravilnih naziva kupaca iz xlsx fajla."""
    try:
        wb = load_workbook(xlsx_path, read_only=True)
        ws = wb.active
        names = []
        for row in ws.iter_rows(min_row=1, max_col=1, values_only=True):
            if row[0]:
                names.append(_fix_cp1250(str(row[0]).strip()))
        wb.close()
        return names
    except Exception:
        return []


_DIACRITICS = str.maketrans('ČĆŽŠĐ', 'CCZSĐ', '')
_DIACRITICS_FULL = str.maketrans('ČĆŽŠĐčćžšđ', 'CCZSDcczsđ', '')


def _normalize_name(name):
    """Normalizuje naziv firme za poređenje."""
    s = name.upper().strip()
    # D.O.O. / D.O.O / D O O → DOO
    s = re.sub(r'D\s*\.\s*O\s*\.\s*O\s*\.?', 'DOO', s)
    s = re.sub(r'\bD\s+O\s+O\b', 'DOO', s)
    # Ukloni tačke, crtice, navodnike
    s = s.replace('.', ' ').replace('-', ' ').replace('"', '').replace("'", '')
    # Višestruki razmaci → jedan
    s = re.sub(r'\s+', ' ', s).strip()
    return s


def _strip_diacritics(name):
    """Uklanja dijakritike (Č→C, Ć→C, Š→S, Ž→Z, Đ→D)."""
    return name.translate(_DIACRITICS_FULL).upper()


def match_kupac_name(extracted_name, known_names):
    """Pronalazi najbolje poklapanje iz liste poznatih kupaca.
    Vraća pravilno ime ako nađe match, inače vraća original."""
    if not extracted_name or not known_names:
        return extracted_name

    norm_extracted = _normalize_name(extracted_name)
    if not norm_extracted:
        return extracted_name

    best_match = None
    best_score = 0

    for known in known_names:
        norm_known = _normalize_name(known)

        # Tačan match nakon normalizacije
        if norm_extracted == norm_known:
            return known

        # Tačan match bez dijakritika (CEVABDZINICA == ĆEVABDŽINICA)
        if _strip_diacritics(norm_extracted) == _strip_diacritics(norm_known):
            return known

        # Jedan sadrži drugi (npr. "ZE TRANS" ⊂ "ZE TRANS DOO")
        if norm_extracted in norm_known or norm_known in norm_extracted:
            overlap = min(len(norm_extracted), len(norm_known))
            max_len = max(len(norm_extracted), len(norm_known))
            score = overlap / max_len if max_len > 0 else 0
            if score > best_score and score >= 0.5:
                best_score = score
                best_match = known
                continue

        # Substring match bez dijakritika
        ascii_ext = _strip_diacritics(norm_extracted)
        ascii_known = _strip_diacritics(norm_known)
        if ascii_ext in ascii_known or ascii_known in ascii_ext:
            overlap = min(len(ascii_ext), len(ascii_known))
            max_len = max(len(ascii_ext), len(ascii_known))
            score = overlap / max_len if max_len > 0 else 0
            if score > best_score and score >= 0.5:
                best_score = score
                best_match = known
                continue

        # Poređenje ključnih riječi (bez DOO, STR, SZR, TR, UR, DD, JP, JU)
        suffixes = {'DOO', 'STR', 'SZR', 'TR', 'UR', 'DD', 'JP', 'JU'}
        words_ext = [w for w in _strip_diacritics(norm_extracted).split() if w not in suffixes]
        words_known = [w for w in _strip_diacritics(norm_known).split() if w not in suffixes]
        if words_ext and words_known:
            common = set(words_ext) & set(words_known)
            total = set(words_ext) | set(words_known)
            score = len(common) / len(total) if total else 0
            if score > best_score and score >= 0.6:
                best_score = score
                best_match = known

    return best_match if best_match else extracted_name


# Učitaj listu kupaca pri importu
_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
KUPCI_NAMES = load_kupci_names(os.path.join(_SCRIPT_DIR, "kupci.xlsx"))

EXTRACTION_PROMPT = """Ovo je račun/faktura. Izvuci polja i vrati kao JSON objekat.

Ključevi MORAJU biti TAČNO ovi (ostavi prazan string "" ako ne postoji):

{
  "BRDOKFAKT": "Broj računa/fakture (npr. 432/10, 9034508513, 600398-1-0126-1)",
  "DATUMF": "Datum izdavanja fakture (format DD.MM.GGGG)",
  "NAZIVPP": "Puni naziv KUPCA - firma KOJOJ je račun izdat (ne firma koja izdaje račun!). Ako iza naziva stoji broj u zagradama npr. 'Novine BH d.o.o (1295)', ZADRŽI taj broj u zagradama!",
  "SJEDISTEPP": "Puna adresa kupca sa poštanskim brojem i mjestom",
  "IDDVPP": "ID broj (JIB) kupca - MORA biti TAČNO 13 cifara i počinjati sa 4. Ako na računu vidiš broj koji nema 13 cifara ili ne počinje sa 4, dodaj vodeću 4 da bude 13 cifara",
  "JIBPUPP": "PDV broj kupca - MORA biti TAČNO 12 cifara. To je isti broj kao ID/JIB ali BEZ vodeće cifre 4. Ako kupac NIJE u PDV sistemu (nema PDV broj na računu), ostavi prazan string",
  "IZNOSNOV": "Iznos BEZ PDV-a (decimalni separator tačka, npr. 155.87)",
  "IZNPDV": "Iznos PDV-a u KM (NE procenat, nego koliko PDV iznosi u novcu, npr. 26.50)",
  "IZNAKFT": "UKUPAN iznos za uplatu SA PDV-om (npr. 182.37)",
  "REF": "PAŽLJIVO PREGLEDAJ CIJELU SLIKU za RUČNO NAPISAN (rukom pisan, hemijskom olovkom) tekst 'REF:' ili 'Ref:' ili 'ref:'. Može biti na BILO KOJEM dijelu papira — na margini, pri vrhu, pri dnu, na poleđini, preko teksta fakture. Iza 'REF:' slijedi iznos (broj). Upiši SAMO taj broj. Npr. ako rukom piše 'REF: 250.00' upiši '250.00'. Ako rukom piše 'REF: 1500' upiši '1500'. Rukopis može biti neuredan! Ako NEMA ručno napisanog 'REF:' teksta, ostavi prazan string ''",
  "OSL": "Provjeri da li na računu postoji tekst o oslobađanju PDV-a. Traži tekst koji sadrži 'oslobođene PDV-a po čl.' ili 'oslobodjene PDV-a po cl.' ili slično. Ako se pominje član 15 ili član 27, upiši '1'. Ako se pominje član 26, upiši '2'. Ako nema takvog teksta, upiši '0'",
  "NAZIV_IZDAVACA": "Naziv firme koja IZDAJE račun (čiji je logo/zaglavlje). Ovo je DOBAVLJAČ, NE kupac!",
  "KUPAC_SIFRA": "PAŽLJIVO TRAŽI broj u ZAGRADAMA odmah iza naziva kupca! Npr. ako piše 'Novine BH d.o.o (1295)' upiši '1295'. Ako piše 'OS MEHMEDALIJA MAK DIZDAR VISOKO (290)' upiši '290'. Ako piše 'JU MJESOVITA SREDNJA SKOLA (196)' upiši '196'. Broj je UVIJEK u oblim zagradama () iza naziva kupca. Ako nema broja u zagradama, ostavi prazan string.",
  "NAZIV_USLUGE": "Pronađi tabelu sa stavkama na računu (kolona 'NAZIV (VRSTA) USLUGE/DOBRA' ili slično). Upiši naziv iz PRVOG reda te tabele. Npr. 'Naša Riječ', 'ZE-DO Eko', 'Oglas' itd. Ako nema tabele sa stavkama, ostavi prazan string."
}

VAŽNO:
- Vrati SAMO čist JSON objekat (NE niz), bez markdown, bez objašnjenja
- Ovo je jedan račun - vrati jedan JSON objekat
- KUPAC je firma na koju glasi račun (piše "Korisnik:", "Kupac:", "Za:" ili slično)
- DOBAVLJAČ/IZDAVAČ je firma čiji je logo/zaglavlje (firma koja ŠALJE račun) - to NIJE kupac!
- Koristi tačku kao decimalni separator (npr. 102.70)
- DATUM: Pažljivo pročitaj GODINU! Trenutna godina je 2025 ili 2026. NE čitaj 2026 kao 2020! Format DD.MM.GGGG
- ID broj (JIB) = 13 cifara, počinje sa 4
- PDV broj = 12 cifara, isti kao JIB bez vodeće 4 (samo firme u PDV sistemu)
- IZNPDV je iznos u KM, NE procenat
- Brojeve prepiši TAČNO
- REF: OBAVEZNO pregledaj CIJELU sliku za RUČNO NAPISAN tekst "REF:" (hemijskom olovkom, rukom). Može biti BILO GDJE na papiru — margine, vrh, dno, dijagonalno, preko teksta. Rukopis može biti neuredan. Upiši SAMO broj iza "REF:". Ako nema ručno napisanog REF, ostavi prazan string
- OSL: Traži tekst "oslobođene/oslobodjene PDV-a po čl. 15/26/27" — čl. 15 ili 27 = "1", čl. 26 = "2", nema = "0"
"""


DNEVNI_HEADERS = [
    "DATUMDOK", "BROJKIFA", "SADRZAJ", "GOTOVINA", "KARTICNO", "DEPOZIT",
]

FISCAL_EXTRACTION_PROMPT = """Na ovoj slici se nalaze fiskalni računi (presjek stanja iz fiskalnog printera).
Može biti od 1 do 5 računa zalijepljenih na jednom papiru.

Za SVAKI račun koji pronađeš, izvuci ova polja:

{
  "DATUMDOK": "Datum dokumenta - nalazi se u vrhu računa, obično ispod 'PRESJEK STANJA' (format DD.MM.GGGG)",
  "BROJKIFA": "",
  "SADRZAJ": "Piši 'DI: ' i samo PRVI broj koji stoji pored DI. Npr. ako na računu piše 'DI: 1532 / 2000', upiši 'DI: 1532'. Ako piše 'DI: 1524 / 2000', upiši 'DI: 1524'. NIKAD ne upisuj '/ 2000' dio!",
  "GOTOVINA": "Iznos pored 'GOTOVINA:' ili 'GOTOVINAR:' — traži u sekciji 'STANJE U KASI:' pri dnu računa. Pažljivo pročitaj svaku cifru! Npr. 75,28 ili 150,89 ili 0,00. Decimalni separator ZAREZ.",
  "KARTICNO": "Iznos pored 'KARTICA:' ili 'KARTICR:' — traži u sekciji 'STANJE U KASI:' pri dnu računa. Pažljivo pročitaj svaku cifru! Npr. 400,46 ili 270,98 ili 0,00. Decimalni separator ZAREZ.",
  "DEPOZIT": "Iznos pored 'DEPOZIT:' — traži u sekciji 'STANJE U KASI:' pri dnu računa. Ako ne postoji, prazan string"
}

VAŽNO:
- Vrati JSON NIZ (array) sa jednim objektom za svaki pronađeni račun
- Ako ima 3 računa na slici, vrati niz od 3 objekta
- BROJKIFA je UVIJEK prazan string ""
- Koristi zarez kao decimalni separator (npr. 75,28)
- DATUM: Pažljivo pročitaj GODINU! Trenutna godina je 2025 ili 2026. NE čitaj 2026 kao 2020! Ako vidiš "2026" to JE 2026, NE 2020. Format: DD.MM.GGGG (bez vremena)
- Ako je vrijednost 0.00 ili 0,00, upiši "0,00"
- Vrati SAMO čist JSON niz, bez markdown, bez objašnjenja
- Pažljivo razdvoji račune - svaki presjek stanja je zaseban račun
- NE miješaj podatke između računa
- GOTOVINA i KARTICA: Čitaj TAČAN iznos cifru po cifru! NE miješaj cifre između računa. Ako tekst iz PDF-a postoji, koristi tekst umjesto slike za precizne brojeve
"""


KUF_HEADERS = [
    "REDBR", "TIPDOK", "BROJFAKT", "DATUMF", "DATUMPF",
    "NAZIVPP", "SJEDISTEPP", "IDPDVPP", "JIBPUPP",
    "IZNBEZPDV", "IZNSAPDV", "IZNPDV", "Moze",
]

KUF_EXTRACTION_PROMPT = """Ovo je ulazni račun/faktura (primljeni od dobavljača). Izvuci polja i vrati kao JSON objekat.

Ključevi MORAJU biti TAČNO ovi (ostavi prazan string "" ako ne postoji):

{
  "BROJFAKT": "Broj računa/fakture (npr. 432/10, 9034508513, 600398-1-0126-1)",
  "DATUMF": "Datum izdavanja fakture (format DD.MM.GGGG)",
  "DATUMPF": "Datum prijema fakture — ako postoji poseban datum prijema/evidentiranja, upiši ga (format DD.MM.GGGG). Ako ne postoji, ostavi prazan string",
  "NAZIVPP": "Puni naziv DOBAVLJAČA — firma KOJA JE IZDALA račun (čiji je logo/zaglavlje). To je firma koja ŠALJE račun, NE firma koja ga prima!",
  "SJEDISTEPP": "Puna adresa dobavljača sa poštanskim brojem i mjestom",
  "IDPDVPP": "ID broj (JIB) dobavljača - MORA biti TAČNO 13 cifara i počinjati sa 4. Ako na računu vidiš broj koji nema 13 cifara ili ne počinje sa 4, dodaj vodeću 4 da bude 13 cifara",
  "JIBPUPP": "PDV broj dobavljača - MORA biti TAČNO 12 cifara. To je isti broj kao ID/JIB ali BEZ vodeće cifre 4. Ako dobavljač NIJE u PDV sistemu (nema PDV broj na računu), ostavi prazan string",
  "IZNBEZPDV": "Iznos BEZ PDV-a (decimalni separator tačka, npr. 155.87)",
  "IZNSAPDV": "UKUPAN iznos za uplatu SA PDV-om (npr. 182.37)",
  "IZNPDV": "Iznos PDV-a u KM (NE procenat, nego koliko PDV iznosi u novcu, npr. 26.50)",
  "Moze": "Provjeri da li se na računu pominje pravo na odbitak ulaznog PDV-a. Ako postoji PDV i nema naznake da se PDV NE može odbiti, upiši '1'. Ako piše da se PDV ne može odbiti ili ako nema PDV-a, upiši '0'"
}

VAŽNO:
- Vrati SAMO čist JSON objekat (NE niz), bez markdown, bez objašnjenja
- Ovo je jedan račun - vrati jedan JSON objekat
- Ovo je ULAZNA faktura — DOBAVLJAČ je firma čiji je logo/zaglavlje (firma koja ŠALJE račun)
- KUPAC/PRIMALAC je firma KOJA PRIMA račun — to NIJE dobavljač!
- Koristi tačku kao decimalni separator (npr. 102.70)
- DATUM: Pažljivo pročitaj GODINU! Trenutna godina je 2025 ili 2026. NE čitaj 2026 kao 2020! Format DD.MM.GGGG
- ID broj (JIB) = 13 cifara, počinje sa 4
- PDV broj = 12 cifara, isti kao JIB bez vodeće 4 (samo firme u PDV sistemu)
- IZNPDV je iznos u KM, NE procenat
- Brojeve prepiši TAČNO
"""


def _chat_completion_with_retry(client, max_retries=5, **kwargs):
    """Poziva chat.completions.create sa retry logikom za 429 rate limit."""
    for attempt in range(max_retries):
        try:
            return client.chat.completions.create(**kwargs)
        except openai.RateLimitError as e:
            if attempt == max_retries - 1:
                raise
            wait = min(2 ** attempt, 30)
            time.sleep(wait)


def process_kuf_pdf(pdf_bytes, filename="", api_key=None):
    """Obrađuje PDF ulazne fakture i vraća dict sa KUF podacima."""
    client = openai.OpenAI(api_key=api_key)

    pdf_text = extract_text_from_bytes(pdf_bytes)

    content = []
    has_text = len(pdf_text) >= MIN_TEXT_LENGTH
    images = pdf_bytes_to_images_base64(pdf_bytes)

    for img in images:
        content.append({
            "type": "image_url",
            "image_url": {"url": f"data:image/png;base64,{img}"},
        })

    if has_text:
        content.append({
            "type": "text",
            "text": f"Pogledaj SLIKU da razumiješ raspored - ko je dobavljač a ko kupac.\n"
                    f"DOBAVLJAČ/IZDAVAČ je firma čiji je logo/zaglavlje (firma koja ŠALJE račun) — TO JE FIRMA ČIJE PODATKE TREBAŠ.\n"
                    f"KUPAC/PRIMALAC je firma na koju glasi račun — to NE trebamo.\n\n"
                    f"Za TAČNE brojeve koristi ovaj tekst iz PDF-a:\n\n"
                    f"---\n{pdf_text}\n---\n\n{KUF_EXTRACTION_PROMPT}",
        })
    else:
        content.append({"type": "text", "text": KUF_EXTRACTION_PROMPT})

    response = _chat_completion_with_retry(
        client,
        model="gpt-4o",
        temperature=0,
        max_tokens=2000,
        messages=[{"role": "user", "content": content}],
    )

    raw = response.choices[0].message.content.strip()

    if raw.startswith("```"):
        raw = raw.split("\n", 1)[1]
        raw = raw.rsplit("```", 1)[0]

    start = raw.find("{")
    end = raw.rfind("}") + 1
    if start >= 0 and end > start:
        raw = raw[start:end]

    data = json.loads(raw)

    # Validacija ID/PDV (ista logika, ali polje se zove IDPDVPP)
    id_broj = str(data.get("IDPDVPP", "")).strip().replace(" ", "")
    if id_broj:
        if len(id_broj) == 12 and not id_broj.startswith("4"):
            id_broj = "4" + id_broj
        if len(id_broj) == 13 and id_broj.startswith("4"):
            data["IDPDVPP"] = id_broj
            pdv = str(data.get("JIBPUPP", "")).strip()
            if not pdv or len(pdv) != 12:
                data["JIBPUPP"] = id_broj[1:]

    # Fiksna polja
    data["REDBR"] = random.randint(1, 10)
    data["TIPDOK"] = "01"

    # Konvertuj brojeve u string sa tačkom kao separatorom
    for key in ["IZNBEZPDV", "IZNSAPDV", "IZNPDV"]:
        val = data.get(key, "")
        if isinstance(val, (int, float)):
            data[key] = f"{val:.2f}"
        elif isinstance(val, str) and val:
            data[key] = val.replace(",", ".")

    # Hardening Moze — mora biti "0" ili "1"
    moze_val = str(data.get("Moze", "0")).strip()
    if moze_val not in ("0", "1"):
        if moze_val.lower() in ("da", "yes", "true"):
            moze_val = "1"
        else:
            moze_val = "0"
    data["Moze"] = moze_val

    return data


def split_pdf_to_pages(pdf_bytes):
    """Razdvaja multi-page PDF na listu single-page PDF bajtova."""
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    pages = []
    for page_num in range(len(doc)):
        single = fitz.open()
        single.insert_pdf(doc, from_page=page_num, to_page=page_num)
        pages.append((page_num + 1, single.tobytes()))
        single.close()
    doc.close()
    return pages


def count_pdf_pages(pdf_bytes):
    """Vraća broj stranica u PDF-u bez čuvanja stranica u memoriji."""
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    n = len(doc)
    doc.close()
    return n


def iter_pdf_pages(pdf_bytes):
    """Generator koji vraća stranice jednu po jednu — ne drži sve u memoriji odjednom."""
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    try:
        for page_num in range(len(doc)):
            single = fitz.open()
            single.insert_pdf(doc, from_page=page_num, to_page=page_num)
            page_bytes = single.tobytes()
            single.close()
            yield (page_num + 1, page_bytes)
    finally:
        doc.close()


def extract_text_from_bytes(pdf_bytes):
    """Izvlači ugrađeni tekst iz PDF bajtova."""
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    text = ""
    for page in doc:
        text += page.get_text()
    doc.close()
    return text.strip()


def pdf_bytes_to_images_base64(pdf_bytes):
    """Konvertuje PDF bajtove u base64 slike."""
    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
        tmp.write(pdf_bytes)
        tmp_path = tmp.name

    try:
        pages = convert_from_path(tmp_path, dpi=150)
        images = []
        for page in pages:
            buffer = BytesIO()
            page.save(buffer, format="PNG")
            img_base64 = base64.b64encode(buffer.getvalue()).decode("utf-8")
            images.append(img_base64)
        return images
    finally:
        os.unlink(tmp_path)


def validate_id_pdv(data):
    """Validacija i korekcija IDDVPP (13 cifara) i JIBPUPP (12 cifara)."""
    id_broj = str(data.get("IDDVPP", "")).strip().replace(" ", "")
    if id_broj:
        if len(id_broj) == 12 and not id_broj.startswith("4"):
            id_broj = "4" + id_broj
        if len(id_broj) == 13 and id_broj.startswith("4"):
            data["IDDVPP"] = id_broj
            pdv = str(data.get("JIBPUPP", "")).strip()
            if not pdv:
                data["JIBPUPP"] = id_broj[1:]
            elif len(pdv) != 12:
                data["JIBPUPP"] = id_broj[1:]
    return data


def process_pdf(pdf_bytes, filename="", api_key=None):
    """Obrađuje PDF i vraća dict sa KIF podacima."""
    client = openai.OpenAI(api_key=api_key)

    pdf_text = extract_text_from_bytes(pdf_bytes)

    content = []
    has_text = len(pdf_text) >= MIN_TEXT_LENGTH

    # UVIJEK šalji sliku — OCR tekst je često pokvarjen i AI treba vidjeti raspored
    images = pdf_bytes_to_images_base64(pdf_bytes)
    for img in images:
        content.append({
            "type": "image_url",
            "image_url": {"url": f"data:image/png;base64,{img}"},
        })

    ref_instruction = (
        "\n\nPOSEBNO VAŽNO — REF polje:\n"
        "Na papiru može biti RUČNO NAPISANO (hemijskom olovkom, rukom) 'REF:' i broj iza toga.\n"
        "Pregledaj CIJELU sliku — margine, uglove, vrh, dno.\n"
        "Ako NEMA ručno napisanog teksta, REF ostavi kao prazan string.\n"
    )

    if has_text:
        content.append({
            "type": "text",
            "text": f"Pogledaj SLIKU da razumiješ raspored - ko je kupac a ko dobavljač.\n"
                    f"DOBAVLJAČ/IZDAVAČ je firma čiji je logo/zaglavlje (firma koja ŠALJE račun).\n"
                    f"KUPAC je firma na koju glasi račun (piše 'Korisnik:', 'Kupac:' ili slično).\n\n"
                    f"Za TAČNE brojeve koristi ovaj tekst iz PDF-a:\n\n"
                    f"---\n{pdf_text}\n---\n\n{EXTRACTION_PROMPT}{ref_instruction}",
        })
    else:
        content.append({"type": "text", "text": f"{EXTRACTION_PROMPT}{ref_instruction}"})

    response = _chat_completion_with_retry(
        client,
        model="gpt-4o",
        temperature=0,
        max_tokens=2000,
        messages=[{"role": "user", "content": content}],
    )

    raw = response.choices[0].message.content.strip()

    if raw.startswith("```"):
        raw = raw.split("\n", 1)[1]
        raw = raw.rsplit("```", 1)[0]

    start = raw.find("{")
    end = raw.rfind("}") + 1
    if start >= 0 and end > start:
        raw = raw[start:end]

    data = json.loads(raw)

    # Dopuni iz poznatih partnera
    full_text = (pdf_text + " " + json.dumps(data)).lower()
    for partner in POZNATI_PARTNERI:
        if any(kw.lower() in full_text for kw in partner["kljucne_rijeci"]):
            if not data.get("NAZIVPP"):
                data["NAZIVPP"] = partner["naziv"]
            if not data.get("IDDVPP"):
                data["IDDVPP"] = partner["id"]
            if not data.get("JIBPUPP"):
                data["JIBPUPP"] = partner["pdv"]
            if not data.get("SJEDISTEPP"):
                data["SJEDISTEPP"] = partner["adresa"]
            break

    # Korekcija naziva iz mape kupaca
    if data.get("NAZIVPP") and KUPCI_NAMES:
        data["NAZIVPP"] = match_kupac_name(data["NAZIVPP"], KUPCI_NAMES)

    # Validacija ID/PDV
    data = validate_id_pdv(data)

    # Fiksna polja
    brdok = str(data.get("BRDOKFAKT", ""))
    redbr_match = re.match(r'(\d+)', brdok)
    data["REDBR"] = int(redbr_match.group(1)) if redbr_match else 0
    data["TIPDOK"] = "01"

    # Konvertuj brojeve u string sa tačkom kao separatorom
    for key in ["IZNAKFT", "IZNOSNOV", "IZNPDV"]:
        val = data.get(key, "")
        if isinstance(val, (int, float)):
            data[key] = f"{val:.2f}"
        elif isinstance(val, str) and val:
            data[key] = val.replace(",", ".")

    # ── Hardening REF ──
    ref_val = str(data.get("REF", "")).strip()
    # Očisti ako je AI vratio "REF:" prefiks ili varijacije (ref, Ref, REF, sa/bez dvotačke)
    ref_val = re.sub(r'^[Rr][Ee][Ff]\s*[:;.\-]?\s*', '', ref_val)
    # Ukloni tekst poput "nema", "nije pronađeno", "prazan" itd.
    if re.match(r'^(nema|nije|prazan|empty|none|n/a|null|ne postoji)', ref_val, re.IGNORECASE):
        ref_val = ""
    # Izvuci broj — dozvoli razmake između cifara (rukopis), tačke, zareze
    if ref_val:
        # Ukloni razmake unutar broja (rukopis: "1 500" → "1500")
        cleaned = re.sub(r'(\d)\s+(\d)', r'\1\2', ref_val)
        ref_match = re.search(r'[\d][,.\d]*[\d]|[\d]', cleaned)
        if ref_match:
            ref_val = ref_match.group().replace(",", ".")
        else:
            ref_val = ""
    # Fallback: provjeri direktno u PDF tekstu ako AI nije pronašao
    if not ref_val and pdf_text:
        ref_text_match = re.search(
            r'[Rr][Ee][Ff]\s*[:;.\-]?\s*([\d][\s,.\d]*[\d]|[\d])',
            pdf_text,
        )
        if ref_text_match:
            ref_val = re.sub(r'\s', '', ref_text_match.group(1)).replace(",", ".")
    data["REF"] = ref_val

    # ── Hardening OSL ──
    osl_val = str(data.get("OSL", "0")).strip()
    # Normaliziraj AI odgovor na 0/1/2
    if osl_val not in ("0", "1", "2"):
        if re.search(r'\b(15|27)\b', osl_val):
            osl_val = "1"
        elif re.search(r'\b26\b', osl_val):
            osl_val = "2"
        else:
            osl_val = "0"
    # Fallback: provjeri direktno u PDF tekstu nezavisno od AI-a
    if pdf_text:
        osl_pattern = re.search(
            r'oslobo[dđ]en[aei]*\s+PDV[\s-]*a\s+po\s+[čc]l[.\s]*(\d+)',
            pdf_text,
            re.IGNORECASE,
        )
        if osl_pattern:
            clan = osl_pattern.group(1)
            if clan in ("15", "27"):
                osl_val = "1"
            elif clan == "26":
                osl_val = "2"
        # Ako nema tog teksta u PDF-u a AI je rekao 0, ostavi 0
        # Ako nema tog teksta a AI je rekao 1 ili 2, provjeri da tekst uopšte postoji
        elif osl_val != "0":
            # AI tvrdi da postoji ali PDF tekst nema — trust AI za skenirane dokumente
            # jer tekst iz slike ne mora biti u pdf_text
            pass
    data["OSL"] = osl_val

    # ── KONTO — samo za fakture izdane od "Naša Riječ" ──
    data["KONTO"] = ""
    # 1) Provjeri direktno u PDF tekstu (pouzdanije od AI-a)
    # OCR često kvari tekst: "NASA R JEC", "NASA RLSEC", "NASA RIJE" itd.
    is_nasa_rijec = bool(
        pdf_text and re.search(r'nasa\s+r', pdf_text, re.IGNORECASE)
    )
    # 2) Fallback: provjeri AI odgovor
    if not is_nasa_rijec:
        izdavac = str(data.get("NAZIV_IZDAVACA", "")).strip()
        is_nasa_rijec = bool(re.search(r'na[sš]a?\s*rije[cč]', izdavac, re.IGNORECASE))

    if is_nasa_rijec:
        # ── OSL za Naša Riječ: 1=Naša Riječ, 2=ZE-DO Eko, 3=ostalo ──
        naziv_usluge = str(data.get("NAZIV_USLUGE", ""))
        # Fallback: ako AI nije izvukao, traži direktno u PDF tekstu
        if not naziv_usluge and pdf_text:
            if re.search(r'ze[\s-]*do', pdf_text, re.IGNORECASE):
                naziv_usluge = "ZE-DO Eko"
            elif re.search(r'na[sš]a\s*rije[cč]', pdf_text, re.IGNORECASE):
                naziv_usluge = "Naša Riječ"
        if re.search(r'ze[\s-]*do', naziv_usluge, re.IGNORECASE):
            data["OSL"] = "2"
        elif re.search(r'na[sš]a?\s*rije[cč]', naziv_usluge, re.IGNORECASE):
            data["OSL"] = "1"
        else:
            data["OSL"] = "3"

        konto_num = None
        # 1) Iz AI polja KUPAC_SIFRA (najdirektnije)
        kupac_sifra = str(data.get("KUPAC_SIFRA", "")).strip()
        if kupac_sifra and kupac_sifra.isdigit():
            konto_num = kupac_sifra
        # 2) Fallback: iz PDF teksta blizu "Kupac"
        if not konto_num and pdf_text:
            kupac_match = re.search(
                r'[Kk]u[pr]a[cr]\s*[:;]?\s*.*?\((\d+)\)',
                pdf_text,
                re.DOTALL,
            )
            if kupac_match:
                konto_num = kupac_match.group(1)
        # 3) Fallback: iz NAZIVPP koji je AI izvukao
        if not konto_num:
            naziv_kupca = str(data.get("NAZIVPP", ""))
            konto_match = re.search(r'\((\d+)\)', naziv_kupca)
            if konto_match:
                konto_num = konto_match.group(1)
        # 4) Fallback: bilo koji broj u zagradama u PDF tekstu
        if not konto_num and pdf_text:
            all_nums = re.findall(r'\((\d{2,5})\)', pdf_text)
            if len(all_nums) == 1:
                konto_num = all_nums[0]
        if konto_num:
            data["KONTO"] = "2112" + konto_num
            data["NAZIVPP"] = re.sub(r'\s*\(\d+\)', '', str(data.get("NAZIVPP", ""))).strip()

    # Ukloni pomoćna polja koja ne idu u tabelu
    data.pop("NAZIV_IZDAVACA", None)
    data.pop("KUPAC_SIFRA", None)
    data.pop("NAZIV_USLUGE", None)

    return data


def _merge_pdf_pages(page_bytes_list):
    """Spaja listu single-page PDF bajtova u jedan PDF."""
    merged = fitz.open()
    for pb in page_bytes_list:
        doc = fitz.open(stream=pb, filetype="pdf")
        merged.insert_pdf(doc)
        doc.close()
    result = merged.tobytes()
    merged.close()
    return result


def _is_incomplete(data):
    """Provjerava da li rezultatu fale ključni iznosi (druga stranica računa)."""
    iznakft = str(data.get("IZNAKFT", "")).strip()
    iznosnov = str(data.get("IZNOSNOV", "")).strip()
    # Ako nema ukupnog iznosa i nema osnove — vjerovatno nepotpun
    return (not iznakft or iznakft == "0" or iznakft == "0.00") and \
           (not iznosnov or iznosnov == "0" or iznosnov == "0.00")


def process_multi_page_pdf(pdf_bytes, filename="", api_key=None):
    """Razdvaja PDF po stranicama. Ako stranica nema iznose, spaja sa sljedećom."""
    pages = split_pdf_to_pages(pdf_bytes)
    results = []
    i = 0
    while i < len(pages):
        page_num, page_bytes = pages[i]
        data = process_pdf(page_bytes, filename=f"{filename} (str. {page_num})", api_key=api_key)

        # Ako fale iznosi i postoji sljedeća stranica — spoji i probaj ponovo
        if _is_incomplete(data) and i + 1 < len(pages):
            next_page_num, next_page_bytes = pages[i + 1]
            merged_bytes = _merge_pdf_pages([page_bytes, next_page_bytes])
            print(f"  [MERGE] Stranica {page_num} nepotpuna, spajam sa {next_page_num}")
            data = process_pdf(merged_bytes, filename=f"{filename} (str. {page_num}-{next_page_num})", api_key=api_key)
            data["_page_num"] = page_num
            data["_page_bytes"] = merged_bytes
            results.append(data)
            i += 2  # Preskoči obje stranice
        else:
            data["_page_num"] = page_num
            data["_page_bytes"] = page_bytes
            results.append(data)
            i += 1

    return results


def process_fiscal_pdf(pdf_bytes, filename="", api_key=None):
    """Obrađuje stranicu sa fiskalnim računima i vraća listu dict-ova."""
    client = openai.OpenAI(api_key=api_key)

    pdf_text = extract_text_from_bytes(pdf_bytes)
    images = pdf_bytes_to_images_base64(pdf_bytes)

    content = []
    for img in images:
        content.append({
            "type": "image_url",
            "image_url": {"url": f"data:image/png;base64,{img}"},
        })

    has_text = len(pdf_text) >= MIN_TEXT_LENGTH
    if has_text:
        content.append({
            "type": "text",
            "text": f"Za TAČNE brojeve koristi ovaj tekst iz PDF-a:\n\n"
                    f"---\n{pdf_text}\n---\n\n{FISCAL_EXTRACTION_PROMPT}",
        })
    else:
        content.append({"type": "text", "text": FISCAL_EXTRACTION_PROMPT})

    response = _chat_completion_with_retry(
        client,
        model="gpt-4o",
        temperature=0,
        max_tokens=4000,
        messages=[{"role": "user", "content": content}],
    )

    raw = response.choices[0].message.content.strip()

    if raw.startswith("```"):
        raw = raw.split("\n", 1)[1]
        raw = raw.rsplit("```", 1)[0]

    # Parsiranje — očekujemo JSON niz
    start = raw.find("[")
    end = raw.rfind("]") + 1
    if start >= 0 and end > start:
        raw = raw[start:end]

    items = json.loads(raw)
    if isinstance(items, dict):
        items = [items]

    results = []
    for data in items:
        data["BROJKIFA"] = ""
        # Konvertuj brojeve u string sa zarezom
        for key in ["GOTOVINA", "KARTICNO", "DEPOZIT"]:
            val = data.get(key, "")
            if isinstance(val, (int, float)):
                data[key] = f"{val:.2f}".replace(".", ",")
            elif isinstance(val, str) and val:
                data[key] = val.replace(".", ",")
        results.append(data)

    return results
